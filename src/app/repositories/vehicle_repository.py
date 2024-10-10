# #vehicle_repository.py
# from datetime import datetime
# from fastapi import HTTPException, status
# from sqlalchemy.orm import Session, joinedload
# from sqlalchemy.orm import aliased
# from sqlalchemy import or_
# #from schemas.vehicle_movement_schemas import VehicleMovementCreate
# from schemas.vehicle_schemas import VehicleCreate, VehicleUpdate #, Update_Fuel_Economy
# from models.all_models import Driver, Employee, FuelPurchase, Location, Site, Trip, TripStage, User, Vehicle, Tenancy, WorkPlan, trip_workplan_association, workplan_site_association,workplan_location_association
# from repositories.base_repository import BaseRepository
# from sqlalchemy.exc import SQLAlchemyError
# from typing import Dict, Optional, List
# import logging
# from logging_helpers import logging_helper
# import csv
# import io
# import openpyxl
# from openpyxl.styles import Font, Alignment
# from openpyxl.utils import get_column_letter
# import xlsxwriter

# class VehicleRepository(BaseRepository[Vehicle, VehicleCreate, VehicleUpdate]):

#     def __init__(self, db_session: Session):
#         super().__init__(Vehicle, db_session)


#     def create_vehicle(self, vehicle_info: VehicleCreate) -> Vehicle:
#         """
#         Creates a new Vehicle ensuring that the licence_plate is unique within the Vehicle table.
#         """

#         licence_plate_existence = self.db_session.query(Vehicle).filter(Vehicle.licence_plate == vehicle_info.licence_plate).first()
#         if licence_plate_existence:
#             logging_helper.log_info(f"The Vehicle {vehicle_info.name} with licence_plate {vehicle_info.licence_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle {vehicle_info.name} with licence_plate {vehicle_info.licence_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
        
#         alternate_plate_existence = self.db_session.query(Vehicle).filter(Vehicle.alternate_plate == vehicle_info.alternate_plate).first()
#         if alternate_plate_existence:
#             logging_helper.log_info(f"The Vehicle {vehicle_info.name} with alternate_plate {vehicle_info.alternate_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle {vehicle_info.name} with alternate_plate {vehicle_info.alternate_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
        
#         tenancy_existence = self.db_session.query(Tenancy).filter(Tenancy.id == vehicle_info.tenancy_id, Tenancy.is_active==True).first()
#         if not tenancy_existence:
#             logging_helper.log_info(f"The Tenancy ID: {vehicle_info.tenancy_id} supplied does not exist")
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Tenancy ID: {vehicle_info.tenancy_id} supplied does not exist")
        
#         try:
#             return self.create(vehicle_info)
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"Database Error during Vehicle creation {err}")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Check the create vehicle route")
    
#     def update_vehicle(self, vehicle_id: int, vehicle_info: VehicleUpdate, tenancy_id: Optional[int] = None) -> Optional[Vehicle]:
#         """
#         Updates an existing Vehicle.
#         """
#         try:
#             if tenancy_id:
#                 db_vehicle = self.get_vehicle_by_id_for_tenancy(
#                     vehicle_id=vehicle_id, tenancy_id=tenancy_id
#                 )
#                 if not db_vehicle:
#                     raise HTTPException(
#                         status_code=status.HTTP_404_NOT_FOUND,
#                         detail=f"The vehicle with id {vehicle_id} does not exist or not in your state",
#                     )
#             else:
#                 db_vehicle = self.get_vehicle_by_id(vehicle_id=vehicle_id)
#                 if not db_vehicle:
#                     raise HTTPException(
#                         status_code=status.HTTP_404_NOT_FOUND,
#                         detail=f"The vehicle with id {vehicle_id} does not exist",)

#             return self.update(db_vehicle, vehicle_info)
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Check the update vehicle route")
        


#     def get_all_active_vehicles(self, skip:int =0, limit:int =100) -> List[Vehicle]:
#         """
#         Retrieves all Vehicles that are active with optional pagination.
#         """
#         try:
#             return self.get_all(skip=skip, limit=limit)
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


#     def get_available_vehicles(self, skip:int =0, limit:int =100, tenancy_id: Optional[int] = None) -> List[Vehicle]:
#         """
#         Retrieves all Vehicles in the records with optional pagination.
#         """
#         try:
#             if tenancy_id:
#                 available_vehicles = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.is_available==True, Vehicle.tenancy_id==tenancy_id).offset(offset=skip).limit(limit=limit).all()
#                 return available_vehicles
#             else:
#                 available_vehicles = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.is_available==True).offset(offset=skip).limit(limit=limit).all()
#                 return available_vehicles
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


#     def get_all_active_vehicle_for_tenancy(
#         self, tenancy_id: int, skip: int = 0, limit: int = 100
#     ) -> List[Vehicle]:
#         """
#         Retrieves all Vehicle that are active with optional pagination.
#         """
#         tenancy_exist = (
#             self.db_session.query(Tenancy).filter(Tenancy.id == tenancy_id).first()
#         )
#         if not tenancy_exist:
#             raise HTTPException(
#                 status_code=status.HTTP_400_BAD_REQUEST,
#                 detail=f"The Tenancy ID {tenancy_id} does not exist",
#             )

#         try:
#             return self.get_all(
#                 skip=skip, limit=limit, tenancy_id=tenancy_id, is_active=True
#             )
#         except Exception as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


#     def get_vehicle_by_id(self, vehicle_id:int) -> Vehicle:
#         """
#         Retrieves a Vehicles by its ID.
#         """
#         try:
#             return self.get_by_id(id=vehicle_id)
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicles") 


#     def get_vehicle_by_id_for_tenancy(self, vehicle_id, tenancy_id) -> Vehicle:
#         """
#         Retrieves a Vehicle by its ID.
#         """
#         try:
#             vehicle = self.get_by_id(id=vehicle_id, tenancy_id=tenancy_id, is_active=True)
#             if vehicle:
#                 return vehicle
#         except Exception as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


#     def get_vehicle_by_licence_plate(self, licence_plate:str, tenancy_id: Optional[int] = None) -> Vehicle :
#         """
#         Retrieves a Vehicles by its licence plate.
#         """
#         try:
#             if tenancy_id:
#                 vehicle = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.tenancy_id==tenancy_id, or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate)).first()
#                 return vehicle
#             else:
#                 vehicle = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate)).first()
#                 return vehicle

#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to retrieve Vehicles stuff {str(err)}") 


#     def delete_vehicle(self, vehicle_id, hard_delete: bool = True, tenancy_id: Optional[int] = None) -> Optional[Vehicle]:
#         """
#         Deletes or soft deletes a Vehicles by ID based on the hard_delete flag.
#         """
#         try:
#             if tenancy_id:
#                 vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
#                 if not vehicle:
#                     raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
#             else:
#                 vehicle = self.get_by_id(id=vehicle_id)
#                 if not vehicle:
#                     raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
#             if hard_delete:
#                 self.db_session.delete(vehicle)
#                 self.db_session.commit()
#                 return f"The Vehicle with ID {vehicle_id} deleted successfully"
#         except SQLAlchemyError as e:
#             logging_helper.log_error(message=f"{str(e)} !!!")
#             raise HTTPException(status_code=500, detail="Failed to delete vehicle")
        

#     def soft_delete_vehicle(self, vehicle_id, tenancy_id: Optional[int] = None):
#         """
#         Soft deletes a vehicle by setting its is_active flag to False.
#         """

#         if tenancy_id:
#             vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
#         else:
#             vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
        
#         if not vehicle.is_active:
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Vehicle {vehicle_id} has already been deactivated")
        
#         try:
#             vehicle.is_active=False
#             self.db_session.commit()
#             return {f"Success"}
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to soft delete Vehicle")

    

#     def restore_vehicle(self, vehicle_id, tenancy_id: Optional[int] = None)->Optional[Vehicle]:
#         """
#         Restores a soft-deleted Vehicle by setting its is_active flag to True.
#         """
#         if tenancy_id:
#             vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
#         else:
#             vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
        
#         if vehicle.is_active:
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Vehicle {vehicle_id} has already been activated")
        
#         try:
#             vehicle.is_active=True
#             self.db_session.commit()
#             return {f"Success"}
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to soft delete Vehicle")



#     def update_fuel_economy(self, licence_plate:str, new_fuel_economy:float, tenancy_id:Optional[int]=None) -> Optional[Vehicle]:
#         """
#         To Update fuel economy of any specified vehicle
#         """

#         vehicle = self.db_session.query(Vehicle).filter(or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate))
        
#         # Check for Tenancy
#         if tenancy_id:
#             vehicle = vehicle.filter(Vehicle.tenancy_id==tenancy_id)
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle with Licence Plate {licence_plate} does not belong to your state.")

#         # Check Vehicle Existence
#         try:
#             vehicle=vehicle.first()
#             if not vehicle:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"The Vehicle with Licence Plate {licence_plate} is not found")
            
#             # Update Vehicle with the new fuel economy
#             vehicle.fuel_economy = new_fuel_economy
#             self.db_session.commit()
#             return vehicle
#         except SQLAlchemyError as err:
#             logging_helper.log_error(message=f"{str(err)} !!!")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed due to {str(err)}")
        

#     def update_vehicle_availability(self, vehicle_id: int, is_available: bool) -> Optional[Vehicle]:
#         """
#         Updates the availability status of a driver.
#         """
#         try:
#             driver = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()

#             if not driver:
#                 raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Driver with ID {vehicle_id} not found.")

#             driver.is_available = is_available
#             self.db_session.commit()

#             return {f"detail: Vehicle Availability updated successfully"}
#         except SQLAlchemyError as err:
#             self.db_session.rollback()
#             logging_helper.log_error(f"Error updating VEhicle availability: {err}")
#             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error updating Vehicle availability: {err}")
        

#     def generate_vehicle_movement_logs_csv(self, start_date: datetime, end_date: datetime, tenancy_ids: Optional[List[int]] = None) -> io.StringIO:
#         logging_helper.log_info("Generating vehicle movement logs CSV")
        
#         try:
#             vehicle_logs = self.get_vehicle_movement_logs(start_date, end_date, tenancy_ids)
#             logging_helper.log_info(f"Retrieved {len(vehicle_logs)} vehicle movement logs")

#             output = io.StringIO()
#             writer = csv.writer(output)

#             # Write header
#             writer.writerow([
#                 "Trip ID", "Vehicle ID", "Vehicle Name", "Vehicle Plate", "Vehicle Type", "Vehicle Capacity",
#                 "Driver ID", "Driver Name", "Driver License", "Activity Date", "Start Mileage", "End Mileage",
#                 "Locations", "Sites", "Employees", "Workplan Lead", "Trip Status", "Activity Title",
#                 "Site Name", "Mileage From", "Mileage To", "Distance Covered", "Arrival Time"
#             ])

#             # Write data
#             for log in vehicle_logs:
#                 if "site_visits" in log and log["site_visits"]:
#                     for site_visit in log["site_visits"]:
#                         writer.writerow([
#                             log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"],
#                             log["vehicle_type"], log["vehicle_capacity"], log["driver_id"], log["driver_name"],
#                             log["driver_license"], log["activity_date"], log["start_mileage"], log["end_mileage"],
#                             log["locations"], log["sites"], log["employees"],
#                             log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "",
#                             log["trip_status"], log["activity_title"],
#                             site_visit["site_name"], site_visit["mileage_from"], site_visit["mileage_to"],
#                             site_visit["distance_covered"], site_visit["arrival_time"]
#                         ])
#                 else:
#                     writer.writerow([
#                         log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"],
#                         log["vehicle_type"], log["vehicle_capacity"], log["driver_id"], log["driver_name"],
#                         log["driver_license"], log["activity_date"], log["start_mileage"], log["end_mileage"],
#                         log["locations"], log["sites"], log["employees"],
#                         log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "",
#                         log["trip_status"], log["activity_title"], "", "", "", "", ""
#                     ])

#             output.seek(0)
#             logging_helper.log_info("Vehicle movement logs CSV generation completed successfully")
#             return output
#         except Exception as e:
#             logging_helper.log_error(f"Error generating vehicle movement logs CSV: {str(e)}")
#             raise HTTPException(status_code=500, detail=f"Error generating vehicle movement logs CSV: {str(e)}")




#     # def get_vehicle_movement_logs(self, start_date: datetime, end_date: datetime, tenancy_ids: List[int] = None) -> List[Dict]:
#     #     """
#     #     Retrieves the vehicle movement logs within a specified period, including associated employees and drivers.
#     #     """
#     #     try:
#     #         # Aliases to avoid duplicate table names
#     #         driver_employee_alias = aliased(Employee)
#     #         workplan_employee_alias = aliased(Employee)
#     #         workplan_lead_alias = aliased(Employee)
            
#     #         # Base query to get trips within the specified period
#     #         query = (
#     #             self.db_session.query(Trip)
#     #             .join(Trip.vehicle)
#     #             .join(Trip.driver)
#     #             .join(Driver.user)
#     #             .join(driver_employee_alias, driver_employee_alias.id == User.employee_id)  # Alias for the driver employee
#     #             .outerjoin(Trip.work_plans)
#     #             .outerjoin(workplan_employee_alias, workplan_employee_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan employees
#     #             .outerjoin(WorkPlan.locations)
#     #             .outerjoin(WorkPlan.sites)
#     #             .outerjoin(workplan_lead_alias, workplan_lead_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan lead
#     #             .options(
#     #                 joinedload(Trip.vehicle),
#     #                 joinedload(Trip.driver).joinedload(Driver.user).joinedload(User.employee),
#     #                 joinedload(Trip.work_plans).joinedload(WorkPlan.employees),
#     #                 joinedload(Trip.work_plans).joinedload(WorkPlan.locations),
#     #                 joinedload(Trip.work_plans).joinedload(WorkPlan.sites),
#     #                 joinedload(Trip.work_plans).joinedload(WorkPlan.activity_lead),
#     #                 joinedload(Trip.trip_stages).joinedload(TripStage.site)
#     #             )
#     #             .filter(Trip.start_time.between(start_date, end_date))
#     #         )

#     #         # Apply tenancy_ids filter if provided
#     #         if tenancy_ids:
#     #             query = query.filter(Trip.tenancy_id.in_(tenancy_ids))

#     #         trips = query.all()

#     #         vehicle_logs = []

#     #         for trip in trips:
#     #             # Retrieve all fuel purchases for the vehicle during the trip dates
#     #             fuel_purchases = (
#     #                 self.db_session.query(FuelPurchase)
#     #                 .filter(
#     #                     FuelPurchase.vehicle_id == trip.vehicle_id,
#     #                     FuelPurchase.purchase_date.between(trip.start_time, trip.end_time if trip.end_time else datetime.utcnow())
#     #                 )
#     #                 .all()
#     #             )

#     #             for workplan in trip.work_plans:
#     #                 trip_info = {
#     #                     "trip_id": trip.id,
#     #                     "vehicle_id": trip.vehicle.id,
#     #                     "vehicle_name": trip.vehicle.name,
#     #                     "vehicle_plate": trip.vehicle.licence_plate,
#     #                     "vehicle_type": trip.vehicle.make,
#     #                     "vehicle_capacity": trip.vehicle.seat_capacity,
#     #                     "driver_id": trip.driver.id,
#     #                     "driver_name": f"{trip.driver.user.employee.first_name} {trip.driver.user.employee.last_name}",  # Full driver name
#     #                     "driver_license": trip.driver.licence_number,
#     #                     "activity_date": trip.start_time,
#     #                     "start_mileage": trip.mileage_at_start,
#     #                     "end_mileage": trip.mileage_at_end,
#     #                     "distance_covered": trip.mileage_at_end - trip.mileage_at_start if trip.mileage_at_end and trip.mileage_at_start else None,
#     #                     "locations": ", ".join([loc.name for loc in workplan.locations]) if workplan else "",
#     #                     "sites": ", ".join([site.name for site in workplan.sites]) if workplan else "",
#     #                     "employees": ", ".join([f"{emp.first_name} {emp.last_name}" for emp in workplan.employees]) if workplan else "",
#     #                     "workplan_lead": {
#     #                         "lead_id": workplan.activity_lead.id,
#     #                         "lead_name": f"{workplan.activity_lead.first_name} {workplan.activity_lead.last_name}"
#     #                     } if workplan and workplan.activity_lead else None,
#     #                     "trip_status": trip.status,
#     #                     "activity_title": workplan.activity_title if workplan else "",
#     #                     "trip_stages": [
#     #                         {
#     #                             "site_name": stage.site.name,
#     #                             "arrival_time": stage.arrival_time,
#     #                             "departure_time": stage.departure_time,
#     #                             "mileage_at_arrival": stage.mileage_at_arrival,
#     #                             "distance_covered": stage.mileage_at_arrival - trip.mileage_at_start if stage.mileage_at_arrival and trip.mileage_at_start else None
#     #                         }
#     #                         for stage in trip.trip_stages
#     #                     ],
#     #                     "start_location": trip.trip_start_location.name if trip.trip_start_location else None,
#     #                     "end_location": trip.trip_end_location.name if trip.trip_end_location else None,
#     #                     "fuel_purchases": [
#     #                         {
#     #                             "amount": purchase.total_cost,
#     #                             "quantity": purchase.quantity,
#     #                             "date": purchase.purchase_date,
#     #                         }
#     #                         for purchase in fuel_purchases
#     #                     ]
#     #                 }

#     #                 vehicle_logs.append(trip_info)

#     #         return vehicle_logs

#     #     except SQLAlchemyError as e:
#     #         logging.error(f"Error retrieving vehicle movement logs: {str(e)}")
#     #         raise HTTPException(
#     #             status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
#     #             detail=f"Failed to retrieve vehicle movement logs: {str(e)}",
#     #         )


#     def get_vehicle_movement_logs(self, start_date: datetime, end_date: datetime, tenancy_ids: List[int] = None) -> List[Dict]:
#         """
#         Retrieves the vehicle movement logs within a specified period, including associated employees, drivers, and fuel purchases.
#         """
#         try:
#             # Aliases to avoid duplicate table names
#             driver_employee_alias = aliased(Employee)
#             workplan_employee_alias = aliased(Employee)
#             workplan_lead_alias = aliased(Employee)
            
#             # Base query to get trips within the specified period
#             query = (
#                 self.db_session.query(Trip)
#                 .join(Trip.vehicle)
#                 .join(Trip.driver)
#                 .join(Driver.user)
#                 # .join(TripStage.trip_id)
#                 .join(driver_employee_alias, driver_employee_alias.id == User.employee_id)  # Alias for the driver employee
#                 .outerjoin(Trip.work_plans)
#                 .outerjoin(workplan_employee_alias, workplan_employee_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan employees
#                 .outerjoin(WorkPlan.locations)
#                 .outerjoin(WorkPlan.sites)
#                 .outerjoin(workplan_lead_alias, workplan_lead_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan lead
#                 .options(
#                     joinedload(Trip.vehicle),
#                     joinedload(Trip.driver).joinedload(Driver.user).joinedload(User.employee),
#                     joinedload(Trip.work_plans).joinedload(WorkPlan.employees),
#                     joinedload(Trip.work_plans).joinedload(WorkPlan.locations),
#                     joinedload(Trip.work_plans).joinedload(WorkPlan.sites),
#                     joinedload(Trip.work_plans).joinedload(WorkPlan.activity_lead),
#                     joinedload(Trip.trip_stages).joinedload(TripStage.site),
#                     joinedload(Trip.vehicle).joinedload(Vehicle.fuel_purchases)
#                 )
#                 .order_by(Trip.id)
#                 .filter(Trip.start_time.between(start_date, end_date))
#             )
#             # Apply tenancy_ids filter if provided
#             if tenancy_ids:
#                 query = query.filter(Trip.tenancy_id.in_(tenancy_ids))

#             trips = query.all()

#             vehicle_logs = []

#             for trip in trips:
#                 for workplan in trip.work_plans:
#                     fuel_purchases = [
#                         {
#                             "quantity": purchase.quantity,
#                             "total_cost": purchase.total_cost,
#                             "purchase_date": purchase.purchase_date
#                         }
#                         for purchase in trip.vehicle.fuel_purchases
#                         if trip.start_time.date() <= purchase.purchase_date <= trip.end_time.date()
#                     ]

#                     stages = trip.trip_stages
#                     stage_routes = []
#                     prev_mileage = trip.mileage_at_start
#                     prev_location = {"from_name":trip.trip_start_location.name, "from_id":trip.trip_start_location.id} if trip.trip_start_location else "Unknown Start Location"
#                     prev_time = trip.start_time

#                     for stage in stages:
#                         distance_covered = stage.mileage_at_arrival - prev_mileage if prev_mileage is not None and stage.mileage_at_arrival is not None  else 0
#                         stage_info = {
#                             "from_name": prev_location.get("from_name"),
#                             "from_id": prev_location.get("from_id"),
#                             "to": stage.site.name,
#                             "to_id":stage.site.id,
#                             "mileage_at_arrival": stage.mileage_at_arrival,
#                             "distance_covered": distance_covered,
#                             "departure_time": stage.departure_time,
#                             "arrival_time": stage.arrival_time
#                         }
#                         stage_routes.append(stage_info)
#                         prev_mileage = stage.mileage_at_arrival
#                         prev_location = {"from_name":stage.site.name, "from_id":stage.site.id}
#                         prev_time = stage.departure_time
                       

#                     if trip.trip_end_location:
#                         final_distance_covered = trip.mileage_at_end - prev_mileage if prev_mileage is not None else 0
#                         prev_location = {"stage_name":stage.site.name, "stage_id":stage.site.id}
#                         final_stage_info = {
#                             "from": prev_location.get("stage_name"),
#                             "from_id": prev_location.get("stage_id"),
#                             "to": trip.trip_end_location.name,
#                             "to_id": trip.trip_end_location.id,
#                             "mileage_at_arrival": trip.mileage_at_end,
#                             "distance_covered": final_distance_covered,
#                             # "arrival_time": prev_time,
#                             "end_location_time": trip.end_time
#                         }
#                         stage_routes.append(final_stage_info)
                        

#                     trip_info = {
#                         "trip_id": trip.id,
#                         "vehicle_id": trip.vehicle.id,
#                         "vehicle_name": trip.vehicle.name,
#                         "vehicle_plate": trip.vehicle.licence_plate,
#                         "vehicle_type": trip.vehicle.make,
#                         "vehicle_capacity": trip.vehicle.seat_capacity,
#                         "driver_id": trip.driver.id,
#                         "driver_name": f"{trip.driver.user.employee.first_name} {trip.driver.user.employee.last_name}",  # Full driver name
#                         "driver_license": trip.driver.licence_number,
#                         "activity_date": trip.start_time,
#                         "start_mileage": trip.mileage_at_start,
#                         "end_mileage": trip.mileage_at_end,
#                         "distance_covered": trip.mileage_at_end - trip.mileage_at_start if trip.mileage_at_end and trip.mileage_at_start else None,
#                         "locations": ", ".join([loc.name for loc in workplan.locations]) if workplan else "",
#                         "sites": ", ".join([site.name for site in workplan.sites]) if workplan else "",
#                         "employees": ", ".join([f"{emp.first_name} {emp.last_name}" for emp in workplan.employees]) if workplan else "",
#                         "workplan_lead": {
#                             "lead_id": workplan.activity_lead.id,
#                             "lead_name": f"{workplan.activity_lead.first_name} {workplan.activity_lead.last_name}"
#                         } if workplan and workplan.activity_lead else None,
#                         "trip_status": trip.status,
#                         "activity_title": workplan.activity_title if workplan else "",
#                         "start_location": trip.trip_start_location.name if trip.trip_start_location else "Unknown Start Location",
#                         "end_location": trip.trip_end_location.name if trip.trip_end_location else "Unknown End Location",
#                         "fuel_purchases": fuel_purchases,
#                         "stage_routes": stage_routes
#                     }

#                     vehicle_logs.append(trip_info)

#             return vehicle_logs

#         except SQLAlchemyError as e:
#             logging.error(f"Error retrieving vehicle movement logs: {str(e)}")
#             raise HTTPException(
#                 status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
#                 detail=f"Failed to retrieve vehicle movement logs: {str(e)}",
#             )


#     def generate_vehicle_movement_logs_xlsx(self, start_date: datetime, end_date: datetime, tenancy_ids: Optional[List[int]] = None) -> io.BytesIO:
#         logging_helper.log_info("Generating vehicle movement logs Excel")

#         try:
#             vehicle_logs = self.get_vehicle_movement_logs(start_date, end_date, tenancy_ids)
#             logging_helper.log_info(f"Retrieved {len(vehicle_logs)} vehicle movement logs")

#             output = io.BytesIO()
#             workbook = xlsxwriter.Workbook(output, {'in_memory': True})
#             worksheet = workbook.add_worksheet()

#             # Write headers with sub-headers
#             headers = [
#                 ("Trip Information", ["Trip ID", "Vehicle ID", "Vehicle Name", "Vehicle Plate", "Vehicle Type", "Vehicle Capacity", 
#                                     "Driver ID", "Driver Name", "Driver License", "Activity Date", "Start Mileage", "End Mileage", 
#                                     "Distance Covered", "Locations", "Sites", "Employees", "Workplan Lead", "Trip Status", "Activity Title",
#                                     "Start Location", "End Location", "Fuel Purchases"]),
#                 ("Trip Stages", ["Site Name", "Mileage At Arrival", "Departure Time", "Arrival Time"])
#             ]

#             row_num = 0
#             col_num = 0
#             for header, sub_headers in headers:
#                 worksheet.merge_range(row_num, col_num, row_num, col_num + len(sub_headers) - 1, header)
#                 for sub_header in sub_headers:
#                     worksheet.write(row_num + 1, col_num, sub_header)
#                     col_num += 1

#             # Write data
#             row_num = 2
#             for log in vehicle_logs:
#                 fuel_purchases = ", ".join(
#                     [f"{purchase['total_cost']} ({purchase['quantity']} L) on {purchase['purchase_date']}" for purchase in log.get("fuel_purchases", [])]
#                 )

#                 if "trip_stages" in log and log["trip_stages"]:
#                     for stage in log["trip_stages"]:
#                         col_num = 0
#                         trip_info = [log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"], log["vehicle_type"],
#                                     log["vehicle_capacity"], log["driver_id"], log["driver_name"], log["driver_license"], log["activity_date"],
#                                     log["start_mileage"], log["end_mileage"], log["distance_covered"], log["locations"], log["sites"], 
#                                     log["employees"], log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "", log["trip_status"],
#                                     log["activity_title"], log["start_location"], log["end_location"], fuel_purchases]
#                         for data in trip_info:
#                             worksheet.write(row_num, col_num, data)
#                             col_num += 1
#                         stage_info = [stage["site_name"], stage["mileage_at_arrival"], stage["departure_time"] if stage["departure_time"] else "", stage["arrival_time"]]
#                         for data in stage_info:
#                             worksheet.write(row_num, col_num, data)
#                             col_num += 1
#                         row_num += 1
#                 else:
#                     col_num = 0
#                     trip_info = [log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"], log["vehicle_type"],
#                                 log["vehicle_capacity"], log["driver_id"], log["driver_name"], log["driver_license"], log["activity_date"],
#                                 log["start_mileage"], log["end_mileage"], log["distance_covered"], log["locations"], log["sites"], 
#                                 log["employees"], log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "", log["trip_status"],
#                                 log["activity_title"], log["start_location"], log["end_location"], fuel_purchases]
#                     for data in trip_info:
#                         worksheet.write(row_num, col_num, data)
#                         col_num += 1
#                     stage_info = ["", "", "", ""]
#                     for data in stage_info:
#                         worksheet.write(row_num, col_num, data)
#                         col_num += 1
#                     row_num += 1

#             workbook.close()
#             output.seek(0)
#             logging_helper.log_info("Vehicle movement logs Excel generation completed successfully")
#             return output
#         except Exception as e:
#             logging_helper.log_error(f"Error generating vehicle movement logs Excel: {str(e)}")
#             raise HTTPException(status_code=500, detail=f"Error generating vehicle movement logs Excel: {str(e)}")





#vehicle_repository.py
from datetime import datetime
from fastapi import HTTPException, status
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.orm import aliased
from sqlalchemy import or_
#from schemas.vehicle_movement_schemas import VehicleMovementCreate
from schemas.vehicle_schemas import VehicleCreate, VehicleUpdate #, Update_Fuel_Economy
from models.all_models import Driver, Employee, FuelPurchase, Location, Site, Trip, TripStage, User, Vehicle, Tenancy, WorkPlan, trip_workplan_association, workplan_site_association,workplan_location_association
from repositories.base_repository import BaseRepository
from sqlalchemy.exc import SQLAlchemyError
from typing import Dict, Optional, List
import logging
from logging_helpers import logging_helper
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import xlsxwriter

class VehicleRepository(BaseRepository[Vehicle, VehicleCreate, VehicleUpdate]):

    def __init__(self, db_session: Session):
        super().__init__(Vehicle, db_session)


    def create_vehicle(self, vehicle_info: VehicleCreate) -> Vehicle:
        """
        Creates a new Vehicle ensuring that the licence_plate is unique within the Vehicle table.
        """

        licence_plate_existence = self.db_session.query(Vehicle).filter(Vehicle.licence_plate == vehicle_info.licence_plate).first()
        if licence_plate_existence:
            logging_helper.log_info(f"The Vehicle {vehicle_info.name} with licence_plate {vehicle_info.licence_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle {vehicle_info.name} with licence_plate {vehicle_info.licence_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
        
        alternate_plate_existence = self.db_session.query(Vehicle).filter(Vehicle.alternate_plate == vehicle_info.alternate_plate).first()
        if alternate_plate_existence:
            logging_helper.log_info(f"The Vehicle {vehicle_info.name} with alternate_plate {vehicle_info.alternate_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle {vehicle_info.name} with alternate_plate {vehicle_info.alternate_plate} already exist for tenancy ID {licence_plate_existence.tenancy_id}")
        
        tenancy_existence = self.db_session.query(Tenancy).filter(Tenancy.id == vehicle_info.tenancy_id, Tenancy.is_active==True).first()
        if not tenancy_existence:
            logging_helper.log_info(f"The Tenancy ID: {vehicle_info.tenancy_id} supplied does not exist")
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Tenancy ID: {vehicle_info.tenancy_id} supplied does not exist")
        
        try:
            return self.create(vehicle_info)
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"Database Error during Vehicle creation {err}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Check the create vehicle route")
    
    def update_vehicle(self, vehicle_id: int, vehicle_info: VehicleUpdate, tenancy_id: Optional[int] = None) -> Optional[Vehicle]:
        """
        Updates an existing Vehicle.
        """
        try:
            if tenancy_id:
                db_vehicle = self.get_vehicle_by_id_for_tenancy(
                    vehicle_id=vehicle_id, tenancy_id=tenancy_id
                )
                if not db_vehicle:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"The vehicle with id {vehicle_id} does not exist or not in your state",
                    )
            else:
                db_vehicle = self.get_vehicle_by_id(vehicle_id=vehicle_id)
                if not db_vehicle:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"The vehicle with id {vehicle_id} does not exist",)

            return self.update(db_vehicle, vehicle_info)
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Check the update vehicle route")
        


    def get_all_active_vehicles(self, skip:int =0, limit:int =100) -> List[Vehicle]:
        """
        Retrieves all Vehicles that are active with optional pagination.
        """
        try:
            return self.get_all(skip=skip, limit=limit)
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


    def get_available_vehicles(self, skip:int =0, limit:int =100, tenancy_id: Optional[int] = None) -> List[Vehicle]:
        """
        Retrieves all Vehicles in the records with optional pagination.
        """
        try:
            if tenancy_id:
                available_vehicles = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.is_available==True, Vehicle.tenancy_id==tenancy_id).offset(offset=skip).limit(limit=limit).all()
                return available_vehicles
            else:
                available_vehicles = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.is_available==True).offset(offset=skip).limit(limit=limit).all()
                return available_vehicles
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


    def get_all_active_vehicle_for_tenancy(
        self, tenancy_id: int, skip: int = 0, limit: int = 100
    ) -> List[Vehicle]:
        """
        Retrieves all Vehicle that are active with optional pagination.
        """
        tenancy_exist = (
            self.db_session.query(Tenancy).filter(Tenancy.id == tenancy_id).first()
        )
        if not tenancy_exist:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"The Tenancy ID {tenancy_id} does not exist",
            )

        try:
            return self.get_all(
                skip=skip, limit=limit, tenancy_id=tenancy_id, is_active=True
            )
        except Exception as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


    def get_vehicle_by_id(self, vehicle_id:int) -> Vehicle:
        """
        Retrieves a Vehicles by its ID.
        """
        try:
            return self.get_by_id(id=vehicle_id)
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicles") 


    def get_vehicle_by_id_for_tenancy(self, vehicle_id, tenancy_id) -> Vehicle:
        """
        Retrieves a Vehicle by its ID.
        """
        try:
            vehicle = self.get_by_id(id=vehicle_id, tenancy_id=tenancy_id, is_active=True)
            if vehicle:
                return vehicle
        except Exception as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to retrieve Vehicle")


    def get_vehicle_by_licence_plate(self, licence_plate:str, tenancy_id: Optional[int] = None) -> Vehicle :
        """
        Retrieves a Vehicles by its licence plate.
        """
        try:
            if tenancy_id:
                vehicle = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, Vehicle.tenancy_id==tenancy_id, or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate)).first()
                return vehicle
            else:
                vehicle = self.db_session.query(Vehicle).filter(Vehicle.is_active==True, or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate)).first()
                return vehicle

        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to retrieve Vehicles stuff {str(err)}") 


    def delete_vehicle(self, vehicle_id, hard_delete: bool = True, tenancy_id: Optional[int] = None) -> Optional[Vehicle]:
        """
        Deletes or soft deletes a Vehicles by ID based on the hard_delete flag.
        """
        try:
            if tenancy_id:
                vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
                if not vehicle:
                    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
            else:
                vehicle = self.get_by_id(id=vehicle_id)
                if not vehicle:
                    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
            if hard_delete:
                self.db_session.delete(vehicle)
                self.db_session.commit()
                return f"The Vehicle with ID {vehicle_id} deleted successfully"
        except SQLAlchemyError as e:
            logging_helper.log_error(message=f"{str(e)} !!!")
            raise HTTPException(status_code=500, detail="Failed to delete vehicle")
        

    def soft_delete_vehicle(self, vehicle_id, tenancy_id: Optional[int] = None):
        """
        Soft deletes a vehicle by setting its is_active flag to False.
        """

        if tenancy_id:
            vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
        else:
            vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
        
        if not vehicle.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Vehicle {vehicle_id} has already been deactivated")
        
        try:
            vehicle.is_active=False
            self.db_session.commit()
            return {f"Success"}
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to soft delete Vehicle")

    

    def restore_vehicle(self, vehicle_id, tenancy_id: Optional[int] = None)->Optional[Vehicle]:
        """
        Restores a soft-deleted Vehicle by setting its is_active flag to True.
        """
        if tenancy_id:
            vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id, Vehicle.tenancy_id==tenancy_id).first()
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist or belong to your state')
        else:
            vehicle = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f'Vehicle {vehicle_id} does not exist')
        
        if vehicle.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Vehicle {vehicle_id} has already been activated")
        
        try:
            vehicle.is_active=True
            self.db_session.commit()
            return {f"Success"}
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to soft delete Vehicle")



    def update_fuel_economy(self, licence_plate:str, new_fuel_economy:float, tenancy_id:Optional[int]=None) -> Optional[Vehicle]:
        """
        To Update fuel economy of any specified vehicle
        """

        vehicle = self.db_session.query(Vehicle).filter(or_(Vehicle.licence_plate==licence_plate, Vehicle.alternate_plate==licence_plate))
        
        # Check for Tenancy
        if tenancy_id:
            vehicle = vehicle.filter(Vehicle.tenancy_id==tenancy_id)
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"The Vehicle with Licence Plate {licence_plate} does not belong to your state.")

        # Check Vehicle Existence
        try:
            vehicle=vehicle.first()
            if not vehicle:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"The Vehicle with Licence Plate {licence_plate} is not found")
            
            # Update Vehicle with the new fuel economy
            vehicle.fuel_economy = new_fuel_economy
            self.db_session.commit()
            return vehicle
        except SQLAlchemyError as err:
            logging_helper.log_error(message=f"{str(err)} !!!")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed due to {str(err)}")
        

    def update_vehicle_availability(self, vehicle_id: int, is_available: bool) -> Optional[Vehicle]:
        """
        Updates the availability status of a driver.
        """
        try:
            driver = self.db_session.query(Vehicle).filter(Vehicle.id == vehicle_id).first()

            if not driver:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Driver with ID {vehicle_id} not found.")

            driver.is_available = is_available
            self.db_session.commit()

            return {f"detail: Vehicle Availability updated successfully"}
        except SQLAlchemyError as err:
            self.db_session.rollback()
            logging_helper.log_error(f"Error updating VEhicle availability: {err}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error updating Vehicle availability: {err}")
        

    def generate_vehicle_movement_logs_csv(self, start_date: datetime, end_date: datetime, tenancy_ids: Optional[List[int]] = None) -> io.StringIO:
        logging_helper.log_info("Generating vehicle movement logs CSV")
        
        try:
            vehicle_logs = self.get_vehicle_movement_logs(start_date, end_date, tenancy_ids)
            logging_helper.log_info(f"Retrieved {len(vehicle_logs)} vehicle movement logs")

            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "Trip ID", "Vehicle ID", "Vehicle Name", "Vehicle Plate", "Vehicle Type", "Vehicle Capacity",
                "Driver ID", "Driver Name", "Driver License", "Activity Date", "Start Mileage", "End Mileage",
                "Locations", "Sites", "Employees", "Workplan Lead", "Trip Status", "Activity Title",
                "Site Name", "Mileage From", "Mileage To", "Distance Covered", "Arrival Time"
            ])

            # Write data
            for log in vehicle_logs:
                if "site_visits" in log and log["site_visits"]:
                    for site_visit in log["site_visits"]:
                        writer.writerow([
                            log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"],
                            log["vehicle_type"], log["vehicle_capacity"], log["driver_id"], log["driver_name"],
                            log["driver_license"], log["activity_date"], log["start_mileage"], log["end_mileage"],
                            log["locations"], log["sites"], log["employees"],
                            log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "",
                            log["trip_status"], log["activity_title"],
                            site_visit["site_name"], site_visit["mileage_from"], site_visit["mileage_to"],
                            site_visit["distance_covered"], site_visit["arrival_time"]
                        ])
                else:
                    writer.writerow([
                        log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"],
                        log["vehicle_type"], log["vehicle_capacity"], log["driver_id"], log["driver_name"],
                        log["driver_license"], log["activity_date"], log["start_mileage"], log["end_mileage"],
                        log["locations"], log["sites"], log["employees"],
                        log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "",
                        log["trip_status"], log["activity_title"], "", "", "", "", ""
                    ])

            output.seek(0)
            logging_helper.log_info("Vehicle movement logs CSV generation completed successfully")
            return output
        except Exception as e:
            logging_helper.log_error(f"Error generating vehicle movement logs CSV: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error generating vehicle movement logs CSV: {str(e)}")




    # def get_vehicle_movement_logs(self, start_date: datetime, end_date: datetime, tenancy_ids: List[int] = None) -> List[Dict]:
    #     """
    #     Retrieves the vehicle movement logs within a specified period, including associated employees and drivers.
    #     """
    #     try:
    #         # Aliases to avoid duplicate table names
    #         driver_employee_alias = aliased(Employee)
    #         workplan_employee_alias = aliased(Employee)
    #         workplan_lead_alias = aliased(Employee)
            
    #         # Base query to get trips within the specified period
    #         query = (
    #             self.db_session.query(Trip)
    #             .join(Trip.vehicle)
    #             .join(Trip.driver)
    #             .join(Driver.user)
    #             .join(driver_employee_alias, driver_employee_alias.id == User.employee_id)  # Alias for the driver employee
    #             .outerjoin(Trip.work_plans)
    #             .outerjoin(workplan_employee_alias, workplan_employee_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan employees
    #             .outerjoin(WorkPlan.locations)
    #             .outerjoin(WorkPlan.sites)
    #             .outerjoin(workplan_lead_alias, workplan_lead_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan lead
    #             .options(
    #                 joinedload(Trip.vehicle),
    #                 joinedload(Trip.driver).joinedload(Driver.user).joinedload(User.employee),
    #                 joinedload(Trip.work_plans).joinedload(WorkPlan.employees),
    #                 joinedload(Trip.work_plans).joinedload(WorkPlan.locations),
    #                 joinedload(Trip.work_plans).joinedload(WorkPlan.sites),
    #                 joinedload(Trip.work_plans).joinedload(WorkPlan.activity_lead),
    #                 joinedload(Trip.trip_stages).joinedload(TripStage.site)
    #             )
    #             .filter(Trip.start_time.between(start_date, end_date))
    #         )

    #         # Apply tenancy_ids filter if provided
    #         if tenancy_ids:
    #             query = query.filter(Trip.tenancy_id.in_(tenancy_ids))

    #         trips = query.all()

    #         vehicle_logs = []

    #         for trip in trips:
    #             # Retrieve all fuel purchases for the vehicle during the trip dates
    #             fuel_purchases = (
    #                 self.db_session.query(FuelPurchase)
    #                 .filter(
    #                     FuelPurchase.vehicle_id == trip.vehicle_id,
    #                     FuelPurchase.purchase_date.between(trip.start_time, trip.end_time if trip.end_time else datetime.utcnow())
    #                 )
    #                 .all()
    #             )

    #             for workplan in trip.work_plans:
    #                 trip_info = {
    #                     "trip_id": trip.id,
    #                     "vehicle_id": trip.vehicle.id,
    #                     "vehicle_name": trip.vehicle.name,
    #                     "vehicle_plate": trip.vehicle.licence_plate,
    #                     "vehicle_type": trip.vehicle.make,
    #                     "vehicle_capacity": trip.vehicle.seat_capacity,
    #                     "driver_id": trip.driver.id,
    #                     "driver_name": f"{trip.driver.user.employee.first_name} {trip.driver.user.employee.last_name}",  # Full driver name
    #                     "driver_license": trip.driver.licence_number,
    #                     "activity_date": trip.start_time,
    #                     "start_mileage": trip.mileage_at_start,
    #                     "end_mileage": trip.mileage_at_end,
    #                     "distance_covered": trip.mileage_at_end - trip.mileage_at_start if trip.mileage_at_end and trip.mileage_at_start else None,
    #                     "locations": ", ".join([loc.name for loc in workplan.locations]) if workplan else "",
    #                     "sites": ", ".join([site.name for site in workplan.sites]) if workplan else "",
    #                     "employees": ", ".join([f"{emp.first_name} {emp.last_name}" for emp in workplan.employees]) if workplan else "",
    #                     "workplan_lead": {
    #                         "lead_id": workplan.activity_lead.id,
    #                         "lead_name": f"{workplan.activity_lead.first_name} {workplan.activity_lead.last_name}"
    #                     } if workplan and workplan.activity_lead else None,
    #                     "trip_status": trip.status,
    #                     "activity_title": workplan.activity_title if workplan else "",
    #                     "trip_stages": [
    #                         {
    #                             "site_name": stage.site.name,
    #                             "arrival_time": stage.arrival_time,
    #                             "departure_time": stage.departure_time,
    #                             "mileage_at_arrival": stage.mileage_at_arrival,
    #                             "distance_covered": stage.mileage_at_arrival - trip.mileage_at_start if stage.mileage_at_arrival and trip.mileage_at_start else None
    #                         }
    #                         for stage in trip.trip_stages
    #                     ],
    #                     "start_location": trip.trip_start_location.name if trip.trip_start_location else None,
    #                     "end_location": trip.trip_end_location.name if trip.trip_end_location else None,
    #                     "fuel_purchases": [
    #                         {
    #                             "amount": purchase.total_cost,
    #                             "quantity": purchase.quantity,
    #                             "date": purchase.purchase_date,
    #                         }
    #                         for purchase in fuel_purchases
    #                     ]
    #                 }

    #                 vehicle_logs.append(trip_info)

    #         return vehicle_logs

    #     except SQLAlchemyError as e:
    #         logging.error(f"Error retrieving vehicle movement logs: {str(e)}")
    #         raise HTTPException(
    #             status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
    #             detail=f"Failed to retrieve vehicle movement logs: {str(e)}",
    #         )


    def get_vehicle_movement_logs(self, start_date: datetime, end_date: datetime, tenancy_ids: List[int] = None) -> List[Dict]:
        """
        Retrieves the vehicle movement logs within a specified period, including associated employees, drivers, and fuel purchases.
        """
        try:
            # Aliases to avoid duplicate table names
            driver_employee_alias = aliased(Employee)
            workplan_employee_alias = aliased(Employee)
            workplan_lead_alias = aliased(Employee)
            
            # Base query to get trips within the specified period
            query = (
                self.db_session.query(Trip)
                .join(Trip.vehicle)
                .join(Trip.driver)
                .join(Driver.user)
                # .join(TripStage.trip_id)
                .join(driver_employee_alias, driver_employee_alias.id == User.employee_id)  # Alias for the driver employee
                .outerjoin(Trip.work_plans)
                .outerjoin(workplan_employee_alias, workplan_employee_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan employees
                .outerjoin(WorkPlan.locations)
                .outerjoin(WorkPlan.sites)
                .outerjoin(workplan_lead_alias, workplan_lead_alias.id == WorkPlan.activity_lead_id)  # Alias for work plan lead
                .options(
                    joinedload(Trip.vehicle),
                    joinedload(Trip.driver).joinedload(Driver.user).joinedload(User.employee),
                    joinedload(Trip.work_plans).joinedload(WorkPlan.employees),
                    joinedload(Trip.work_plans).joinedload(WorkPlan.locations),
                    joinedload(Trip.work_plans).joinedload(WorkPlan.sites),
                    joinedload(Trip.work_plans).joinedload(WorkPlan.activity_lead),
                    joinedload(Trip.trip_stages).joinedload(TripStage.site),
                    joinedload(Trip.vehicle).joinedload(Vehicle.fuel_purchases)
                )
                .order_by(Trip.id)
                .filter(Trip.start_time.between(start_date, end_date))
            )

            # Apply tenancy_ids filter if provided
            if tenancy_ids:
                query = query.filter(Trip.tenancy_id.in_(tenancy_ids))

            trips = query.all()

            vehicle_logs = []

            for trip in trips:
                for workplan in trip.work_plans:
                    fuel_purchases = [
                        {
                            "quantity": purchase.quantity,
                            "total_cost": purchase.total_cost,
                            "purchase_date": purchase.purchase_date
                        }
                        for purchase in trip.vehicle.fuel_purchases
                        if trip.start_time.date() <= purchase.purchase_date <= trip.end_time.date()
                    ]

                    stages = trip.trip_stages
                    stage_routes = []
                    prev_mileage = trip.mileage_at_start
                    prev_location = {"from_name":trip.trip_start_location.name, "from_id":trip.trip_start_location.id} if trip.trip_start_location else "Unknown Start Location"
                    prev_time = trip.start_time

                    for stage in stages:
                        distance_covered = stage.mileage_at_arrival - prev_mileage if prev_mileage is not None and stage.mileage_at_arrival is not None  else 0
                        stage_info = {
                            "from_name": prev_location.get("from_name"),
                            "from_id": prev_location.get("from_id"),
                            "to": stage.site.name,
                            "to_id":stage.site.id,
                            "mileage_at_arrival": stage.mileage_at_arrival,
                            "distance_covered": distance_covered,
                            "departure_time": stage.departure_time,
                            "arrival_time": stage.arrival_time,
                            "stage_id": stage.id
                        }
                        stage_routes.append(stage_info)
                        prev_mileage = stage.mileage_at_arrival
                        prev_location = {"from_name":stage.site.name, "from_id":stage.site.id}
                        prev_time = stage.departure_time
                       

                    if trip.trip_end_location:
                        final_distance_covered = trip.mileage_at_end - prev_mileage if prev_mileage is not None else 0
                        prev_location = {"stage_name":stage.site.name, "stage_id":stage.site.id}
                        final_stage_info = {
                            "from": prev_location.get("stage_name"),
                            "from_id": prev_location.get("stage_id"),
                            "to": trip.trip_end_location.name,
                            "to_id": trip.trip_end_location.id,
                            "mileage_at_arrival": trip.mileage_at_end,
                            "distance_covered": final_distance_covered,
                            # "arrival_time": prev_time,
                            "end_location_time": trip.end_time
                        }
                        stage_routes.append(final_stage_info)
                        

                    trip_info = {
                        "trip_id": trip.id,
                        "vehicle_id": trip.vehicle.id,
                        "vehicle_name": trip.vehicle.name,
                        "vehicle_plate": trip.vehicle.licence_plate,
                        "vehicle_type": trip.vehicle.make,
                        "vehicle_capacity": trip.vehicle.seat_capacity,
                        "driver_id": trip.driver.id,
                        "driver_name": f"{trip.driver.user.employee.first_name} {trip.driver.user.employee.last_name}",  # Full driver name
                        "driver_license": trip.driver.licence_number,
                        "activity_date": trip.start_time,
                        "start_mileage": trip.mileage_at_start,
                        "end_mileage": trip.mileage_at_end,
                        "distance_covered": trip.mileage_at_end - trip.mileage_at_start if trip.mileage_at_end and trip.mileage_at_start else None,
                        "locations": ", ".join([loc.name for loc in workplan.locations]) if workplan else "",
                        "sites": ", ".join([site.name for site in workplan.sites]) if workplan else "",
                        "employees": ", ".join([f"{emp.first_name} {emp.last_name}" for emp in workplan.employees]) if workplan else "",
                        "workplan_lead": {
                            "lead_id": workplan.activity_lead.id,
                            "lead_name": f"{workplan.activity_lead.first_name} {workplan.activity_lead.last_name}"
                        } if workplan and workplan.activity_lead else None,
                        "trip_status": trip.status,
                        "activity_title": workplan.activity_title if workplan else "",
                        "start_location": trip.trip_start_location.name if trip.trip_start_location else "Unknown Start Location",
                        "end_location": trip.trip_end_location.name if trip.trip_end_location else "Unknown End Location",
                        "fuel_purchases": fuel_purchases,
                        "stage_routes": stage_routes
                    }

                    vehicle_logs.append(trip_info)

            return vehicle_logs

        except SQLAlchemyError as e:
            logging.error(f"Error retrieving vehicle movement logs: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to retrieve vehicle movement logs: {str(e)}",
            )


    def generate_vehicle_movement_logs_xlsx(self, start_date: datetime, end_date: datetime, tenancy_ids: Optional[List[int]] = None) -> io.BytesIO:
        logging_helper.log_info("Generating vehicle movement logs Excel")

        try:
            vehicle_logs = self.get_vehicle_movement_logs(start_date, end_date, tenancy_ids)
            logging_helper.log_info(f"Retrieved {len(vehicle_logs)} vehicle movement logs")

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()

            # Write headers with sub-headers
            headers = [
                ("Trip Information", ["Trip ID", "Vehicle ID", "Vehicle Name", "Vehicle Plate", "Vehicle Type", "Vehicle Capacity", 
                                    "Driver ID", "Driver Name", "Driver License", "Activity Date", "Start Mileage", "End Mileage", 
                                    "Distance Covered", "Locations", "Sites", "Employees", "Workplan Lead", "Trip Status", "Activity Title",
                                    "Start Location", "End Location", "Fuel Purchases"]),
                ("Trip Stages", ["Site Name", "Mileage At Arrival", "Departure Time", "Arrival Time"])
            ]

            row_num = 0
            col_num = 0
            for header, sub_headers in headers:
                worksheet.merge_range(row_num, col_num, row_num, col_num + len(sub_headers) - 1, header)
                for sub_header in sub_headers:
                    worksheet.write(row_num + 1, col_num, sub_header)
                    col_num += 1

            # Write data
            row_num = 2
            for log in vehicle_logs:
                fuel_purchases = ", ".join(
                    [f"{purchase['total_cost']} ({purchase['quantity']} L) on {purchase['purchase_date']}" for purchase in log.get("fuel_purchases", [])]
                )

                if "trip_stages" in log and log["trip_stages"]:
                    for stage in log["trip_stages"]:
                        col_num = 0
                        trip_info = [log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"], log["vehicle_type"],
                                    log["vehicle_capacity"], log["driver_id"], log["driver_name"], log["driver_license"], log["activity_date"],
                                    log["start_mileage"], log["end_mileage"], log["distance_covered"], log["locations"], log["sites"], 
                                    log["employees"], log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "", log["trip_status"],
                                    log["activity_title"], log["start_location"], log["end_location"], fuel_purchases]
                        for data in trip_info:
                            worksheet.write(row_num, col_num, data)
                            col_num += 1
                        stage_info = [stage["site_name"], stage["mileage_at_arrival"], stage["departure_time"] if stage["departure_time"] else "", stage["arrival_time"]]
                        for data in stage_info:
                            worksheet.write(row_num, col_num, data)
                            col_num += 1
                        row_num += 1
                else:
                    col_num = 0
                    trip_info = [log["trip_id"], log["vehicle_id"], log["vehicle_name"], log["vehicle_plate"], log["vehicle_type"],
                                log["vehicle_capacity"], log["driver_id"], log["driver_name"], log["driver_license"], log["activity_date"],
                                log["start_mileage"], log["end_mileage"], log["distance_covered"], log["locations"], log["sites"], 
                                log["employees"], log["workplan_lead"]["lead_name"] if log["workplan_lead"] else "", log["trip_status"],
                                log["activity_title"], log["start_location"], log["end_location"], fuel_purchases]
                    for data in trip_info:
                        worksheet.write(row_num, col_num, data)
                        col_num += 1
                    stage_info = ["", "", "", ""]
                    for data in stage_info:
                        worksheet.write(row_num, col_num, data)
                        col_num += 1
                    row_num += 1

            workbook.close()
            output.seek(0)
            logging_helper.log_info("Vehicle movement logs Excel generation completed successfully")
            return output
        except Exception as e:
            logging_helper.log_error(f"Error generating vehicle movement logs Excel: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error generating vehicle movement logs Excel: {str(e)}")
